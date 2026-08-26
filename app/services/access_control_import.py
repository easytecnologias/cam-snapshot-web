"""Importacao de alunos por planilha.

A escola ja tem a lista pronta -- digitar 500 alunos na tela nao e opcao. Aqui a
planilha e lida, conferida e so entao gravada, com pre-visualizacao obrigatoria:
o usuario ve quantos serao criados, quantos atualizados e quais linhas foram
recusadas, antes de qualquer escrita.

A matricula e a chave: linha com matricula que ja existe ATUALIZA a pessoa em vez
de criar outra. Isso torna a importacao repetivel -- a escola pode mandar a lista
corrigida quantas vezes quiser sem duplicar ninguem.

O que a planilha NAO traz e a foto. Importar cria o cadastro e o telefone do
responsavel, mas o rosto continua vindo da controladora ou de upload individual.
"""
from __future__ import annotations

import io
import re
import unicodedata
from typing import Any, Dict, List, Tuple

# Cada campo aceita varios nomes de coluna porque cada secretaria escreve do seu
# jeito. A comparacao ignora acento, caixa e espaco extra.
COLUNAS: Dict[str, Tuple[str, ...]] = {
    "enrollment_code": ("matricula", "matriculas", "codigo", "cod", "ra", "registro", "numero"),
    "full_name": ("nome", "nome completo", "aluno", "nome do aluno", "estudante"),
    "guardian_phone": ("telefone", "celular", "whatsapp", "contato", "fone",
                       "telefone do responsavel", "celular do responsavel"),
    "guardian_name": ("responsavel", "nome do responsavel", "mae", "pai", "filiacao"),
    "class_name": ("turma", "classe", "serie", "ano", "sala"),
    "document_id": ("cpf", "documento", "rg", "doc"),
    "controller_user_id": ("id controladora", "id da controladora", "id catraca",
                           "usuario controladora", "id do equipamento"),
}

OBRIGATORIAS = {"enrollment_code": "matricula", "full_name": "nome"}


def _norm(valor: Any) -> str:
    txt = str(valor or "").strip().lower()
    txt = unicodedata.normalize("NFKD", txt)
    txt = "".join(c for c in txt if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", txt)


def _texto(valor: Any) -> str:
    if valor is None:
        return ""
    # openpyxl devolve numero para celula numerica: matricula 1577 viria 1577.0
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def _telefone(valor: Any) -> str:
    """Normaliza para o formato que a Cloud API espera (digitos com DDI).

    Telefone torto passa despercebido e vira notificacao que nunca chega -- e o
    pai quem descobre. Por isso o que nao tem cara de numero brasileiro valido e
    recusado na leitura, em vez de gravado torto.
    """
    digitos = re.sub(r"\D", "", _texto(valor))
    if not digitos:
        return ""
    if digitos.startswith("55") and len(digitos) in (12, 13):
        return digitos
    if len(digitos) in (10, 11):
        return "55" + digitos
    return ""


def _mapear_colunas(cabecalho: List[Any]) -> Dict[str, int]:
    """Descobre em que coluna esta cada campo, pelos nomes aceitos."""
    achadas: Dict[str, int] = {}
    for i, bruto in enumerate(cabecalho):
        nome = _norm(bruto)
        if not nome:
            continue
        for campo, apelidos in COLUNAS.items():
            if campo in achadas:
                continue
            if nome in apelidos or any(nome.startswith(a) for a in apelidos):
                achadas[campo] = i
                break
    return achadas


def ler_planilha(conteudo: bytes, nome_arquivo: str = "") -> List[List[Any]]:
    """Devolve as linhas cruas, aceitando XLSX ou CSV."""
    parece_xlsx = conteudo[:2] == b"PK"
    if not parece_xlsx:
        texto = None
        for codec in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                texto = conteudo.decode(codec)
                break
            except UnicodeDecodeError:
                continue
        if texto is None:
            raise ValueError("Nao foi possivel ler o arquivo: codificacao desconhecida.")
        import csv as _csv

        # Excel brasileiro exporta CSV com ponto e virgula
        separador = ";" if texto.count(";") > texto.count(",") else ","
        return [linha for linha in _csv.reader(io.StringIO(texto), delimiter=separador)]

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    aba = wb[wb.sheetnames[0]]
    return [list(linha) for linha in aba.iter_rows(values_only=True)]


def analisar_planilha(conteudo: bytes, *, site: str = "", nome_arquivo: str = "") -> Dict[str, Any]:
    """Le a planilha e diz o que aconteceria, sem gravar nada."""
    from app.services.access_control_store import list_people

    linhas = [l for l in ler_planilha(conteudo, nome_arquivo) if any(_texto(c) for c in l)]
    if not linhas:
        raise ValueError("A planilha esta vazia.")

    colunas = _mapear_colunas(linhas[0])
    faltando = [rotulo for campo, rotulo in OBRIGATORIAS.items() if campo not in colunas]
    if faltando:
        vistas = ", ".join(str(c) for c in linhas[0] if _texto(c)) or "nenhuma"
        raise ValueError(
            "A planilha precisa das colunas: %s. Colunas encontradas: %s."
            % (", ".join(faltando), vistas)
        )

    existentes: Dict[str, Any] = {}
    for pessoa in (list_people() or []):
        matricula = _texto(pessoa.get("enrollment_code"))
        if matricula:
            existentes[matricula] = pessoa

    criar: List[Dict[str, Any]] = []
    atualizar: List[Dict[str, Any]] = []
    recusados: List[Dict[str, Any]] = []
    vistas_no_arquivo: Dict[str, int] = {}

    for numero, linha in enumerate(linhas[1:], start=2):

        def campo(nome: str) -> str:
            i = colunas.get(nome)
            return _texto(linha[i]) if i is not None and i < len(linha) else ""

        matricula = campo("enrollment_code")
        nome = campo("full_name")
        if not matricula or not nome:
            recusados.append({"linha": numero, "matricula": matricula, "nome": nome,
                              "motivo": "matricula ou nome em branco"})
            continue
        if matricula in vistas_no_arquivo:
            recusados.append({"linha": numero, "matricula": matricula, "nome": nome,
                              "motivo": "matricula repetida na planilha (linha %d)"
                                        % vistas_no_arquivo[matricula]})
            continue
        vistas_no_arquivo[matricula] = numero

        telefone_bruto = campo("guardian_phone")
        telefone = _telefone(telefone_bruto)
        if telefone_bruto and not telefone:
            recusados.append({"linha": numero, "matricula": matricula, "nome": nome,
                              "motivo": "telefone invalido: %s" % telefone_bruto})
            continue

        registro = {
            "linha": numero,
            "enrollment_code": matricula,
            "full_name": nome,
            "guardian_phone": telefone,
            "guardian_name": campo("guardian_name"),
            "class_name": campo("class_name"),
            "document_id": campo("document_id"),
            "controller_user_id": re.sub(r"\D", "", campo("controller_user_id")),
            "site": site,
            "sem_telefone": not telefone,
        }
        if matricula in existentes:
            registro["id_existente"] = existentes[matricula]["id"]
            atualizar.append(registro)
        else:
            criar.append(registro)

    return {
        "ok": True,
        "colunas_reconhecidas": sorted(colunas.keys()),
        "total_linhas": len(linhas) - 1,
        "criar": criar,
        "atualizar": atualizar,
        "recusados": recusados,
        "sem_telefone": sum(1 for r in criar + atualizar if r["sem_telefone"]),
    }


def aplicar_planilha(conteudo: bytes, *, site: str = "", nome_arquivo: str = "") -> Dict[str, Any]:
    """Grava o que a analise aprovou. Linha recusada nao entra."""
    from app.services.access_control_store import save_person

    analise = analisar_planilha(conteudo, site=site, nome_arquivo=nome_arquivo)
    criados = 0
    atualizados = 0
    falhas: List[Dict[str, Any]] = []

    for registro in analise["criar"] + analise["atualizar"]:
        payload = {k: v for k, v in registro.items()
                   if k not in ("linha", "sem_telefone", "id_existente")}
        if registro.get("id_existente"):
            payload["id"] = registro["id_existente"]
        if not payload.get("site"):
            # sem site escolhido na tela, nao apaga o que a pessoa ja tinha
            payload.pop("site", None)
        payload["person_type"] = "student"
        try:
            save_person(payload)
            if registro.get("id_existente"):
                atualizados += 1
            else:
                criados += 1
        except Exception as exc:
            falhas.append({"linha": registro["linha"],
                           "matricula": registro["enrollment_code"],
                           "motivo": str(exc)})

    return {
        "ok": True,
        "criados": criados,
        "atualizados": atualizados,
        "recusados": analise["recusados"],
        "falhas": falhas,
        "sem_telefone": analise["sem_telefone"],
    }
